import openpyxl

# Read data in aCell of Excel File

def read_data_of_cell(path,row,column):

    worksheet=openpyxl.load_workbook(path)
    sheet=worksheet.active
    cell_data=sheet.cell(row=row,column=column)
    return cell_data.value

def read_data_of_row(filepath):
    credentials=[]
    worksheet=openpyxl.load_workbook(filepath)
    sheet=worksheet.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        credentials.append({
            "username":row[1],
            "password":row[2]
        })

    return credentials
